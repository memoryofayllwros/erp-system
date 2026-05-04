import logging
import os
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)


async def create_worker_info_excel(
    workers_data: List[Dict[str, Any]], output_filename: Optional[str] = None
) -> str:

    if not output_filename:
        output_filename = f"worker_info.xlsx"

    columns = [
        "Work Type",
        "Staff No.",
        "Payee Name",
        "Chinese Name",
        "English Name",
        "Occupation",
        "1st Phone",
        "2nd Phone"
    ]

    all_rows = []
    for worker in workers_data:
        # Get phone numbers
        first_phone = worker.get("1st_phone", "")
        second_phone = worker.get("2nd_phone", "")
        
        # Format first phone number
        formatted_phone_1 = ""
        if first_phone and str(first_phone).strip():
            clean_mobile = "".join(filter(str.isdigit, str(first_phone)))
            if len(clean_mobile) >= 4:
                formatted_phone_1 = f"{clean_mobile[:4]} {clean_mobile[4:]}"
            else:
                formatted_phone_1 = clean_mobile
        
        # Format second phone number
        formatted_phone_2 = ""
        if second_phone and str(second_phone).strip():
            clean_mobile = "".join(filter(str.isdigit, str(second_phone)))
            if len(clean_mobile) >= 4:
                formatted_phone_2 = f"{clean_mobile[:4]} {clean_mobile[4:]}"
            else:
                formatted_phone_2 = clean_mobile

        row_data = {
            "Work Type": worker.get("work_type", ""),
            "Staff No.": worker.get("staff_no", ""),
            "Payee Name": worker.get("payee_name", ""),
            "Chinese Name": worker.get("chinese_name", ""),
            "English Name": worker.get("english_name", ""),
            "Occupation": worker.get("occupation", ""),
            "1st Phone": formatted_phone_1,
            "2nd Phone": formatted_phone_2,
        }
        all_rows.append(row_data)

    df = pd.DataFrame(all_rows, columns=columns)

    # Create Excel file in memory
    from io import BytesIO
    
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Worker Information", index=False)
        
        workbook = writer.book
        worksheet = writer.sheets["Worker Information"]
        
        _format_worksheet(worksheet, len(columns))
    
    # Get bytes from buffer
    excel_buffer.seek(0)
    excel_bytes = excel_buffer.read()
    
    # Save to GridFS
    from infrastructure.database.database_connection import get_grid_fs
    from src.utils.datetime_standarization_helpers import get_this_moment
    
    grid_fs = await get_grid_fs()
    
    metadata = {
        "upload_date": get_this_moment().isoformat(),
        "file_type": "worker_list",
        "worker_count": len(workers_data)
    }
    
    file_id = await grid_fs.upload_from_stream(
        filename=output_filename,
        source=BytesIO(excel_bytes),
        metadata=metadata
    )
    
    logger.info(f"✅ Created and uploaded worker information Excel file: {output_filename} with ID: {file_id}")
    return str(file_id)


def _format_worksheet(worksheet, num_columns: int):
    """

    Args:
        worksheet: The openpyxl worksheet object
        num_columns: Number of columns in the worksheet
    """

    header_font = Font(
        name="Songti SC", bold=True, size=12, color="000000"
    )  # Black text
    header_fill = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    cell_font = Font(name="Songti SC", size=10, color="000000")  # Black text
    cell_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, num_columns + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


async def insert_worker_data(output_filename: Optional[str] = None) -> str:

    from src.models.user_model import User

    try:

        users = await User.find(User.deleted_at == None).to_list()

        workers_data = []
        for user in users:
            # Get mobile numbers - handle both list and string formats
            mobile_list = user.mobile if isinstance(user.mobile, list) else [user.mobile] if user.mobile else []
            phone_1 = mobile_list[0] if len(mobile_list) > 0 else ""
            phone_2 = mobile_list[1] if len(mobile_list) > 1 else ""

            # Get work type - convert enum to string if needed
            work_type = user.work_type.value if hasattr(user.work_type, 'value') else str(user.work_type) if user.work_type else ""

            worker_data = {
                "work_type": work_type,
                "staff_no": user.staff_no or "",
                "payee_name": user.payee_name or "",
                "chinese_name": user.chinese_name or "",
                "english_name": user.english_name or "",
                "occupation": user.occupation or "",
                "1st_phone": phone_1,
                "2nd_phone": phone_2,
            }
            workers_data.append(worker_data)

        return await create_worker_info_excel(workers_data, output_filename)

    except Exception as e:
        logger.error(f"❌ Error creating worker Excel file: {str(e)}")
        raise


async def update_existing_worker_excel(
    file_path: str, workers_data: List[Dict[str, Any]]
) -> str:
    """
    Args:
        file_path: Path to the existing Excel file
        workers_data: List of dictionaries containing worker information

    Returns:
        Path to the updated Excel file
    """

    try:
        if os.path.exists(file_path):
            df_existing = pd.read_excel(file_path, sheet_name="Worker Information")
        else:
            columns = [
                "工人編號",
                "英文名",
                "中文名",
                "身份證號碼",
                "出生日期 (d/m/yy)",
                "銀行代碼",
                "銀行戶口號碼",
                "電話",
                "住址",
                "註冊證號碼",
                "註冊證到期日",
                "平安卡號碼",
                "平安卡到期日",
                "合約薪金",
                "工種分項(1)",
                "工種分項(2)",
                "工種分項(3)",
                "工種分項(4)",
                "工種分項(5)",
                "出糧判頭",
            ]
            df_existing = pd.DataFrame(columns=columns)

        new_rows = []
        for worker in workers_data:
            mobile = worker.get("mobile", "")
            logging.info(f"Original mobile: {mobile}")

            if mobile and str(mobile).strip():
                clean_mobile = "".join(filter(str.isdigit, str(mobile)))
                logging.info(f"Cleaned mobile: {clean_mobile}")

                if len(clean_mobile) >= 4:
                    formatted_mobile = f"{clean_mobile[:4]} {clean_mobile[4:]}"
                    logging.info(f"Formatted mobile: {formatted_mobile}")
                else:
                    formatted_mobile = clean_mobile
                    logging.info(f"Short number, no formatting: {formatted_mobile}")
            else:
                formatted_mobile = mobile
                logging.info(f"Empty mobile, using original: {formatted_mobile}")

            phone_number = f"{formatted_mobile}"
            row_data = {
                "工人編號": worker.get("worker_no", ""),
                "英文名": worker.get("english_name", ""),
                "中文名": worker.get("chinese_name", ""),
                "身份證號碼": worker.get("national_id_no", ""),
                "出生日期 (d/m/yy)": worker.get("dob", ""),
                "銀行代碼": worker.get("bank_code", ""),
                "銀行戶口號碼": worker.get("bank_account", ""),
                "電話": phone_number,
                "住址": worker.get("address", ""),
                "註冊證號碼": worker.get("registration_number", ""),
                "註冊證到期日": worker.get("registration_expiry", ""),
                "平安卡號碼": worker.get("safety_card_number", ""),
                "平安卡到期日": worker.get("safety_card_expiry", ""),
                "合約薪金": worker.get("base_daily_salary", ""),
                "工種分項(1)": worker.get("work_type_1", ""),
                "工種分項(2)": worker.get("work_type_2", ""),
                "工種分項(3)": worker.get("work_type_3", ""),
                "工種分項(4)": worker.get("work_type_4", ""),
                "工種分項(5)": worker.get("work_type_5", ""),
                "出糧判頭": worker.get("payroll_contractor", ""),
            }
            new_rows.append(row_data)

        df_new = pd.DataFrame(new_rows)
        df_updated = pd.concat([df_existing, df_new], ignore_index=True)

        # Save updated file
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_updated.to_excel(writer, sheet_name="Worker Information", index=False)

            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets["Worker Information"]
            _format_worksheet(worksheet, len(df_updated.columns))

        logger.info(f"✅ Updated worker Excel file: {file_path}")
        return file_path

    except Exception as e:
        logger.error(f"❌ Error updating worker Excel file: {str(e)}")
        raise


async def generate_worker_info_xlsx(workers_data: List[Dict[str, Any]]) -> str:
    return await create_worker_info_excel(workers_data)
