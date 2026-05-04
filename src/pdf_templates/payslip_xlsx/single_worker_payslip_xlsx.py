from __future__ import annotations

import calendar
import logging
import os
from decimal import Decimal
from typing import List, Optional, Sequence, Tuple

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage

from src.pdf_templates.payslip_xlsx.payslip_helpers import (
    PayslipDayEntry, _each_worker_income_calculation, _safe_currency)


# Helper: add an outline border to a rectangular cell range without removing existing inner borders
def _add_outline_border(
    ws,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    *,
    style: str = "thin",
) -> None:
    outline_side = Side(style=style)

    # Top edge
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=outline_side,
            bottom=cell.border.bottom,
        )

    # Bottom edge
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=end_row, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=outline_side,
        )

    # Left edge
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=start_col)
        cell.border = Border(
            left=outline_side,
            right=cell.border.right,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )

    # Right edge
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=end_col)
        cell.border = Border(
            left=cell.border.left,
            right=outline_side,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )


virpluz_logo_path = "./app/Assets/Image/virpluz_logo.png"


def _prepare_logo_image(max_height: int = 52) -> str | None:
    """Return a path to a preprocessed company logo image suitable for Excel.

    Strategy:
    - Prefer an existing PNG; fall back to JPG/other known assets.
    - Resize with high-quality resampling to a maximum height while keeping aspect ratio.
    - Save to a stable temporary PNG path that openpyxl can read from.
    """
    candidates = [
        "./app/Assets/Image/virpluz_logo.png",
        "./app/Assets/Image/virpluz_logo.jpg",
        "./app/Assets/Image/logo.png",
        "./app/Assets/Image/resized_logo_temp.png",
    ]

    source_path: str | None = None
    for path in candidates:
        if os.path.exists(path):
            source_path = path
            break

    if not source_path:
        return None

    try:
        with PILImage.open(source_path) as im:
            im = im.convert("RGBA")
            width, height = im.size
            if height > max_height and height > 0:
                scale = max_height / float(height)
                new_size = (int(round(width * scale)), int(round(height * scale)))
                im = im.resize(new_size, resample=PILImage.LANCZOS)

            # Save to a consistent temporary PNG path inside assets
            out_path = "./app/Assets/Image/resized_logo_temp.png"
            im.save(out_path, format="PNG", optimize=True)
            return out_path
    except Exception:
        # If preprocessing fails, just return the original path
        return source_path


# Internal helper to draw a payslip into an existing worksheet
def _populate_payslip_sheet(
    ws,
    *,
    employee_chinese_name: str,
    employee_english_name: str,
    employee_code: str,
    id_number: str,
    country_code: str,
    phone: str,
    address: str,
    banker: str,
    bank_account_no: str,
    project_location: str,
    year: int,
    month: int,
    normalized_entries: Sequence[PayslipDayEntry],
    base_daily_salary: Decimal,
    over_65: bool,
    additional_bonus: Optional[Decimal] = None,
    hourly_rate: Decimal,
):
    # Prefer PNG, but fall back to JPG or default logos if not present

    logging.info(f"OT hours rate: {hourly_rate}")
    income_data = _each_worker_income_calculation(
        year=year,
        month=month,
        day_entries=normalized_entries,
        base_daily_salary=base_daily_salary,
        over_65=over_65,
        hourly_rate=hourly_rate,
    )

    base_daily_salary = income_data["base_daily_salary"]  # 日薪工資 with ot_salary
    ot_salary_total = income_data["ot_salary"]
    employee_mpf_amount = income_data["employee_mpf_amount"]
    employer_mpf_amount = income_data["employer_mpf_amount"]
    real_salary = income_data["real_salary"]
    total_payment = income_data["total_payment"]

    # Find and preprocess a suitable logo image
    logo_path = _prepare_logo_image(max_height=52) or virpluz_logo_path

    # Column widths to resemble the provided layout
    column_widths = {
        "A": 8.5,  # Day
        "B": 10.5,  # AM
        "C": 8.5,  # PM
        "D": 8.5,  # OT
        "E": 8.5,  # Day (right)
        "F": 6.0,  # AM
        "G": 8.5,  # PM
        "H": 8.5,  # OT
        "I": 6.0,  # Spacer
        "J": 6.0,  # Spacer
        "K": 8.5,  # Notes/labels
        "L": 18.0,  # Values
        "M": 15.0,  # Values
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Some reusable styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # Place letterhead image across the top rows if found
    img = ExcelImage(logo_path)

    # Scale image to max height 52 while preserving aspect ratio
    original_height = getattr(img, "height", None)
    original_width = getattr(img, "width", None)
    if original_height and original_width and original_height > 52:
        scale = 52 / float(original_height)
        img.height = int(round(original_height * scale))
        img.width = int(round(original_width * scale))

    # Merge A1:A2 so the logo occupies two rows in column A
    ws.merge_cells("A1:A2")
    # Cap column A width at 8.5
    ws.column_dimensions["A"].width = 8.5
    # Keep reasonable row heights (do not exceed requested cap)
    ws.row_dimensions[1].height = min(52, max(ws.row_dimensions[1].height or 18, 18))
    ws.row_dimensions[2].height = min(52, max(ws.row_dimensions[2].height or 18, 18))

    img.anchor = "A1"
    ws.add_image(img)

    current_row = 4

    ws["B2"] = "Virpluz Limited"
    ws["B2"].font = Font(size=16, bold=True)
    ws["B2"].alignment = left

    ws["B3"] = "Flat 13, 6/F, 99 Commons, 99 Pui To Road, Tuen Mun, N.T., Hong Kong"
    ws["B3"].font = Font(size=8)
    ws["B3"].alignment = left

    ws["B4"] = "Tel: 3614 5094; Fax: 3615 1392; Email: info@Virpluz.hk"
    ws["B4"].font = Font(size=8)
    ws["B4"].alignment = left

    # Employee info block placed explicitly in cells next to logo
    # Row 5
    ws["A5"] = "姓名"
    ws["A5"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B5"] = employee_chinese_name
    ws["D5"] = employee_english_name
    ws["F5"] = "Staff No."
    ws["F5"].alignment = Alignment(horizontal="right", vertical="center")
    ws["G5"] = employee_code

    # Row 6
    ws["A6"] = "身份證"
    ws["A6"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B6"] = id_number

    # Row 7
    ws["A7"] = "地址"
    ws["A7"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("B7:D7")
    ws["B7"] = address

    # Row 8
    ws["A8"] = "電話"
    ws["A8"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B8"] = phone
    ws["E8"] = "Bank Acc No."
    ws["E8"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("F8:G8")
    ws["F8"] = bank_account_no

    # Row 9
    ws["A9"] = "工作地點"
    ws["A9"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("B9:D9")
    ws["B9"] = project_location

    # Row 10
    ws["A10"] = "時期"
    ws["A10"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B10"] = f"{year}年{month}月"

    current_row = 12

    # Attendance grid headers
    header_row = current_row
    ws.cell(row=header_row, column=1, value=None).border = thin_border
    ws.cell(row=header_row, column=2, value="上午").alignment = center
    ws.cell(row=header_row, column=3, value="下午").alignment = center
    ws.cell(row=header_row, column=4, value="OT").alignment = center

    ws.cell(row=header_row, column=5, value=None).border = thin_border
    ws.cell(row=header_row, column=6, value="上午").alignment = center
    ws.cell(row=header_row, column=7, value="下午").alignment = center
    ws.cell(row=header_row, column=8, value="OT").alignment = center

    for col in (1, 2, 3, 4, 5, 6, 7, 8):
        c = ws.cell(row=header_row, column=col)
        c.alignment = center
        c.border = thin_border

    # Fill attendance rows
    check_mark = 1
    num_days = calendar.monthrange(year, month)[1]
    left_rows = min(16, num_days)

    grid_start_row = header_row + 1
    row_pointer = grid_start_row
    for idx, entry in enumerate(normalized_entries, start=1):
        if idx > num_days:
            break

        if idx <= 16:
            ws.cell(row=row_pointer, column=1, value=idx).alignment = center
            ws.cell(row=row_pointer, column=1).border = thin_border
            ws.cell(
                row=row_pointer, column=2, value=check_mark if entry.am_worked else None
            ).alignment = center
            ws.cell(
                row=row_pointer, column=3, value=check_mark if entry.pm_worked else None
            ).alignment = center
            ws.cell(
                row=row_pointer, column=4, value=(entry.ot_hours or None)
            ).alignment = center
            for col in (2, 3, 4):
                ws.cell(row=row_pointer, column=col).border = thin_border
            row_pointer += 1
        else:
            # Right half
            right_row = grid_start_row + (idx - 17)
            ws.cell(row=right_row, column=5, value=idx).alignment = center
            ws.cell(row=right_row, column=5).border = thin_border
            ws.cell(
                row=right_row, column=6, value=check_mark if entry.am_worked else None
            ).alignment = center
            ws.cell(
                row=right_row, column=7, value=check_mark if entry.pm_worked else None
            ).alignment = center
            ws.cell(
                row=right_row, column=8, value=(entry.ot_hours or None)
            ).alignment = center
            for col in (6, 7, 8):
                ws.cell(row=right_row, column=col).border = thin_border

    # 'Others' label should be at E28
    ws.cell(row=28, column=5, value="Others").alignment = center
    ws.cell(row=28, column=5).border = thin_border

    # Compute work counts
    full_days = 0.0
    half_days = 0.0
    total_ot_hours = Decimal(0.0)
    for idx, entry in enumerate(normalized_entries, start=1):
        if idx > num_days:
            break
        worked_halves = int(entry.am_worked) + int(entry.pm_worked)
        if worked_halves >= 2:
            full_days += 1
        elif worked_halves == 1:
            half_days += 1
        total_ot_hours += Decimal(entry.ot_hours or 0)

    # Place quick summary counts at fixed rows: 工作日數 at A30, 加班 at A32
    ws.cell(row=30, column=1, value="工作日數/day").alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=30, column=2, value="=SUM(B13:C28,F13:G28)/2").alignment = center

    ws.cell(row=31, column=1, value="加班/hour").alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=31, column=2, value="=SUM(D13:D28,H13:H28)").alignment = center

    bonus_row = 32
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws.merge_cells(start_row=bonus_row, start_column=5, end_row=bonus_row, end_column=6)
    ws.cell(row=bonus_row, column=5, value="獎金").alignment = (
        center  # "獎金" in cell E32:F32
    )

    ws.merge_cells(start_row=bonus_row, start_column=7, end_row=bonus_row, end_column=8)
    ws.cell(row=bonus_row, column=7, value="扣鄭家祥").alignment = (
        center  # "扣鄭家祥" in cell G32:H32
    )
    ws.cell(row=bonus_row, column=7).fill = highlight

    # Left block labels/values - exact positions
    base_daily_salary_row = 33
    ws.merge_cells(
        start_row=base_daily_salary_row,
        start_column=1,
        end_row=base_daily_salary_row,
        end_column=2,
    )
    ws.cell(row=base_daily_salary_row, column=1, value="日薪").alignment = (
        center  # A33:B33
    )

    ws.merge_cells(
        start_row=base_daily_salary_row,
        start_column=3,
        end_row=base_daily_salary_row,
        end_column=4,
    )
    ws.cell(
        row=base_daily_salary_row, column=3, value=_safe_currency(base_daily_salary)
    ).alignment = center  # C33:D33 #日薪工資

    ws.merge_cells(start_row=45, start_column=1, end_row=45, end_column=2)
    ws.cell(row=45, column=1, value="日薪 + 獎金").alignment = Alignment(
        horizontal="right", vertical="center"
    )  # "日薪 + 獎金" in cell A45:B45
    ws.cell(
        row=45,
        column=3,
        value=base_daily_salary
        + (additional_bonus if additional_bonus else Decimal(0)),
    ).alignment = center  # standard日薪工資 + 獎金 base_daily_salary + bonus

    ws.merge_cells(start_row=33, start_column=7, end_row=33, end_column=8)
    ws.cell(
        row=33, column=7, value=additional_bonus if additional_bonus else 0
    ).alignment = center

    # Amount row for daily wage total
    amount_row = 34
    ws.merge_cells(
        start_row=amount_row, start_column=1, end_row=amount_row, end_column=2
    )
    ws.cell(row=amount_row, column=1, value="金額").alignment = center  # A34:B34
    ws.merge_cells(
        start_row=amount_row, start_column=3, end_row=amount_row, end_column=4
    )
    ws.cell(row=amount_row, column=3, value="=ROUNDUP(B30*C33,1)").alignment = (
        center  # C34:D34
    )

    # Right block: 加班費 (use monthly total OT hours from attendance) on the same row with amount_row
    ws.merge_cells(
        start_row=amount_row, start_column=5, end_row=amount_row, end_column=6
    )
    ws.cell(row=amount_row, column=5, value="加班費").alignment = center  # E35:F35

    ws.merge_cells(
        start_row=amount_row, start_column=7, end_row=amount_row, end_column=8
    )
    ws.cell(row=amount_row, column=7, value=ot_salary_total).alignment = (
        center  # G35:H35
    )

    # Right block: second 金額
    second_amount_row = 35
    ws.merge_cells(
        start_row=second_amount_row,
        start_column=5,
        end_row=second_amount_row,
        end_column=6,
    )
    ws.cell(row=second_amount_row, column=5, value="金額").alignment = center

    ws.merge_cells(
        start_row=second_amount_row,
        start_column=7,
        end_row=second_amount_row,
        end_column=8,
    )
    ws.cell(
        row=second_amount_row, column=7, value="=ROUNDUP(B30*G33+G34,1)"
    ).alignment = center  # G35:H35

    # MPF rows (employee + employer). We subtract only employee for take-home.
    mpf_row_left = 37
    ws.merge_cells(
        start_row=mpf_row_left, start_column=1, end_row=mpf_row_left, end_column=2
    )
    ws.cell(row=mpf_row_left, column=1, value="僱員強積金").alignment = (
        center  # A37:B37
    )

    employee_mpf_amount = income_data["employee_mpf_amount"]

    ws.merge_cells(
        start_row=mpf_row_left, start_column=3, end_row=mpf_row_left, end_column=4
    )
    ws.cell(row=mpf_row_left, column=3, value=employee_mpf_amount).alignment = (
        center  # C37:D37
    )

    # Employer MPF on the right block (keep within E-F for label and amount)
    ws.merge_cells(
        start_row=mpf_row_left, start_column=5, end_row=mpf_row_left, end_column=6
    )
    ws.cell(row=mpf_row_left, column=5, value="僱主強積金").alignment = (
        center  # E37:F37
    )

    employer_mpf_amount = income_data["employer_mpf_amount"]

    ws.merge_cells(
        start_row=mpf_row_left, start_column=7, end_row=mpf_row_left, end_column=8
    )
    ws.cell(row=mpf_row_left, column=7, value=employer_mpf_amount).alignment = (
        center  # G37:H37
    )

    # Real payable row (left block bottom)
    real_salary_row = 38
    ws.merge_cells(
        start_row=real_salary_row, start_column=1, end_row=real_salary_row, end_column=2
    )
    ws.cell(row=real_salary_row, column=1, value="實支金額").alignment = center  # A38

    ws.merge_cells(
        start_row=real_salary_row, start_column=3, end_row=real_salary_row, end_column=4
    )
    ws.cell(row=real_salary_row, column=3, value=real_salary).alignment = center  # C38

    # Total amount row (right block bottom) 總金額
    total_amount_row = 38
    ws.merge_cells(
        start_row=total_amount_row,
        start_column=5,
        end_row=total_amount_row,
        end_column=6,
    )
    ws.cell(row=total_amount_row, column=5, value="總金額").alignment = center  # E38

    ws.merge_cells(
        start_row=total_amount_row,
        start_column=7,
        end_row=total_amount_row,
        end_column=8,
    )
    ws.cell(row=total_amount_row, column=7, value=total_payment).alignment = (
        center  # F38
    )

    # Workday counts block on the bottom left to match screenshot
    full_day_counts_row = 41
    ws.merge_cells(
        start_row=full_day_counts_row,
        start_column=1,
        end_row=full_day_counts_row,
        end_column=2,
    )
    ws.cell(row=full_day_counts_row, column=1, value="全日工作日數").alignment = (
        Alignment(horizontal="right", vertical="center")
    )  # "全日工作日數" in cell A41:B41

    ws.cell(row=full_day_counts_row, column=3, value=int(full_days)).alignment = center
    for c in (1, 2, 3):
        ws.cell(row=full_day_counts_row, column=c).border = thin_border

    ws.merge_cells(
        start_row=full_day_counts_row + 1,
        start_column=1,
        end_row=full_day_counts_row + 1,
        end_column=2,
    )
    ws.cell(row=full_day_counts_row + 1, column=1, value="半日工作日數").alignment = (
        Alignment(horizontal="right", vertical="center")
    )  # "半日工作日數" in cell A42:B42
    ws.cell(row=full_day_counts_row + 1, column=3, value=int(half_days)).alignment = (
        center
    )
    for c in (1, 2, 3):
        ws.cell(row=full_day_counts_row + 1, column=c).border = thin_border

    # Signature and date placeholders on the right
    confirm_row = full_day_counts_row
    ws.cell(row=confirm_row, column=5, value="確認簽收").alignment = (
        center  # "確認簽收" in cell E41
    )
    display_name = (
        employee_chinese_name or employee_english_name
    )  # "確認簽收" in cell F41
    ws.cell(row=confirm_row, column=6, value=display_name).alignment = center
    for c in (5, 6):
        ws.cell(row=confirm_row, column=c).border = thin_border

    ws.cell(row=46, column=5, value="日期:").alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=46, column=6, value=None).alignment = center
    for c in (5, 6):
        ws.cell(row=46, column=c).border = thin_border

    # Default number formats for currency-ish cells
    currency_cells: List[Tuple[int, int]] = []
    # Left block currency
    for r, c in [
        (base_daily_salary_row, 3),  # C33:D33 merged start
        (amount_row, 3),  # C34:D34 merged start
        (mpf_row_left, 3),  # C37:D37 merged start
        (real_salary_row, 3),  # C38:D38 merged start
        (45, 3),  # C45
    ]:
        currency_cells.append((r, c))

    # Right block currency
    currency_cells.append((second_amount_row, 7))  # G35:H35 merged start

    for r, c in currency_cells:
        try:
            ws.cell(row=r, column=c).number_format = "#,##0.00"
        except Exception:
            pass

    # Add outline border around A33:H38 (salary + OT + MPF + totals block)
    _add_outline_border(
        ws, start_row=33, start_col=1, end_row=38, end_col=8, style="thin"
    )
