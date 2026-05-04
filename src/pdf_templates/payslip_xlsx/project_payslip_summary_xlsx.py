from __future__ import annotations

import os
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)

from src.pdf_templates.payslip_xlsx.payslip_helpers import \
    _each_worker_income_calculation
# Reuse the day-entry structure and helpers from the payslip generator
from src.pdf_templates.payslip_xlsx.single_worker_payslip_xlsx import \
    PayslipDayEntry


@dataclass
class ProjectSummaryItem:
    """Input record for a single worker in the summary.

    - day_entries can be a sequence of PayslipDayEntry or dicts with keys
      like {day, am/am_worked/morning, pm/pm_worked/afternoon, ot/ot_hours}.
    - allowances are sequences of (label, amount).
    """

    employee_chinese_name: str
    employee_english_name: str
    employee_code: str = ""
    banker: str = ""
    bank_account_no: str = ""
    id_number: str = ""
    over_65: bool = False
    base_daily_salary: Decimal = 0.0
    day_entries: Sequence[PayslipDayEntry] | Sequence[Dict[str, Any]] = ()
    allowances: Optional[Sequence[Tuple[str, Decimal]]] = None
    # Optional precomputed fields (if caller prefers to pass them)
    employee_mpf: Optional[Decimal] = None
    employer_mpf: Optional[Decimal] = None
    ot_hours: Optional[Decimal] = None
    ot_salary: Optional[Decimal] = None
    sheet_name: str = ""  # Name of the individual employee's sheet for hyperlink
    bonus: Optional[Decimal] = None  # Bonus amount for the employee


def add_summary_sheet_to_workbook(
    wb: Workbook,
    *,
    year: int,
    month: int,
    items: Sequence[ProjectSummaryItem] | Sequence[Dict[str, Any]],
    sheet_title: str = "Summary",
    project_location: str,
):
    """Add a Summary sheet into an existing workbook."""

    # Normalize input items into SummaryItem objects
    norm_items: List[ProjectSummaryItem] = []
    for obj in items:
        if isinstance(obj, ProjectSummaryItem):
            norm_items.append(obj)
        elif isinstance(obj, dict):
            norm_items.append(
                ProjectSummaryItem(
                    employee_chinese_name=str(
                        obj.get("employee_chinese_name") or obj.get("name") or ""
                    ),
                    employee_english_name=str(
                        obj.get("employee_english_name")
                        or obj.get("english_name")
                        or ""
                    ),
                    employee_code=str(
                        obj.get("employee_code") or obj.get("code") or ""
                    ),
                    banker=str(obj.get("banker") or ""),
                    bank_account_no=str(
                        obj.get("bank_account_no") or obj.get("bank_no") or ""
                    ),
                    id_number=str(obj.get("id_number") or ""),
                    base_daily_salary=Decimal(obj.get("base_daily_salary") or 0.0),
                    over_65=bool(obj.get("over_65", False)),
                    day_entries=(obj.get("day_entries") or ()),
                    allowances=(obj.get("allowances") or []),
                    employee_mpf=(
                        Decimal(obj.get("employee_mpf"))
                        if obj.get("employee_mpf") is not None
                        else None
                    ),
                    employer_mpf=(
                        Decimal(obj.get("employer_mpf"))
                        if obj.get("employer_mpf") is not None
                        else None
                    ),
                    ot_hours=(
                        Decimal(obj.get("ot_hours"))
                        if obj.get("ot_hours") is not None
                        else None
                    ),
                    ot_salary=(
                        Decimal(obj.get("ot_salary"))
                        if obj.get("ot_salary") is not None
                        else None
                    ),
                    sheet_name=str(obj.get("sheet_name") or ""),
                    bonus=(
                        Decimal(obj.get("bonus"))
                        if obj.get("bonus") is not None
                        else None
                    ),
                )
            )
        else:
            raise TypeError("Unsupported summary item type. Use SummaryItem or dict")

    ws = wb.create_sheet(title=sheet_title)

    # Title and period
    ws["B1"] = "Site:"
    ws["B1"].font = Font(size=12)
    ws["C1"] = project_location
    ws["C1"].font = Font(size=12)

    ws["B2"] = "時期"
    ws["B2"].font = Font(size=12)
    ws["C2"] = f"{year}年{month}月"
    ws["C2"].font = Font(size=12)

    # Column widths to fit the grid
    widths = {
        "A": 6.0,  # 序號
        "B": 8.0,  # chinese name
        "C": 16.0,  # english name
        "D": 20.0,  # HKID
        "E": 10.0,  # banker
        "F": 8.0,  # account no
        "G": 8.0,  # basic salary
        "H": 10.0,  # basic bonus
        "I": 12.0,  # full day
        "J": 12.0,  # half day
        "K": 12.0,  # checking
        "L": 12.0,  # working days
        "M": 14.0,  # OT hours
        "N": 12.0,  # daily salary
        "O": 12.0,  # employee MPF
        "P": 12.0,  # employer MPF
        "Q": 14.0,  # real salary 實收金額 = daily salary - employee MPF
        "R": 12.0,  # bonus
        "S": 12.0,  # total payment 總金額 = daily salary + bonus + employer MPF
        "T": 12.0,  # subcontractor 判頭
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    header_row = 4  # row height is 26
    headers = [
        "CODE",
        "工人",
        "英文名",
        "HKID",
        "Banker",
        "Acc No.",
        "Basic Salary",
        "Basic Bonus",
        "全日",
        "半日",
        "Checking",
        "工作日數",
        "加班",
        "日薪金額",
        "僱員強積金",
        "僱主強積金",
        "實收金額",
        "獎金",
        "總金額",
        "判頭",
    ]
    for idx, text in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=text)
        c.font = Font(bold=True)
        c.alignment = center
        c.border = thin_border

    # Freeze below header
    ws.freeze_panes = ws["A5"]

    # Freeze right side
    ws.freeze_panes = ws["C1"]

    # Rows
    start_row = header_row + 1
    currency_cols = {7, 8, 14, 15, 16, 17, 18, 19}  # monetary columns
    integer_cols = {9, 10}  # full, half
    decimal_cols = {12, 13}  # equivalent days, OT hours

    totals = {
        "full_days": Decimal(0),
        "half_days": Decimal(0),
        "equivalent_days": Decimal(0),
        "ot_hours": Decimal(0),
        "ot_salary": Decimal(0),
        "daily_salary": Decimal(0),
        "base_daily_salary": Decimal(0),
        "employee_mpf_amount": Decimal(0),
        "employer_mpf_amount": Decimal(0),
        "real_salary": Decimal(0),
        "total_payment": Decimal(0),
        "bonus": Decimal(0),
    }

    for row_idx, item in enumerate(norm_items, start=start_row):
        calc = _each_worker_income_calculation(
            year=year,
            month=month,
            day_entries=item.day_entries,
            base_daily_salary=item.base_daily_salary,
            over_65=item.over_65,
            hourly_rate=item.hourly_rate,
        )

        totals["full_days"] += calc["full_days"]
        totals["half_days"] += calc["half_days"]
        totals["equivalent_days"] += calc["equivalent_days"]
        totals["ot_hours"] += calc["ot_hours"]
        totals["ot_salary"] += calc["ot_salary"]
        totals["daily_salary"] += calc["daily_salary"]
        totals["employee_mpf_amount"] += calc["employee_mpf_amount"]
        totals["employer_mpf_amount"] += calc["employer_mpf_amount"]
        totals["real_salary"] += calc["real_salary"]
        totals["total_payment"] += calc["total_payment"]
        totals["base_daily_salary"] += calc["base_daily_salary"]

        values = [
            item.employee_code,  # CODE
            item.employee_chinese_name,  # 工人
            item.employee_english_name,  # 英文名
            item.id_number,  # HKID
            item.banker,  # Banker
            item.bank_account_no,  # Acc No.
            calc["base_daily_salary"],  # Basic Salary
            0,  # Basic Bonus (allowances if any)
            int(calc["full_days"]),  # 全日
            int(calc["half_days"]),  # 半日
            0,  # Checking (not used)
            calc["equivalent_days"],  # 工作日數
            calc["ot_hours"],  # 加班
            calc["daily_salary"],  # 日薪金額
            calc["employee_mpf_amount"],  # 僱員強積金
            calc["employer_mpf_amount"],  # 僱主強積金
            calc["real_salary"],  # 實收金額
            0,  # 獎金
            calc["total_payment"],  # 總金額
            "",  # 判頭
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center if col_idx != 2 else left
            cell.border = thin_border

            # Create hyperlink for employee name (column B) if sheet_name is available
            if col_idx == 2 and item.sheet_name:  # Column B is employee_chinese_name
                try:
                    cell.hyperlink = f"#{item.sheet_name}!A1"
                    # Apply hyperlink styling (blue color and underline)
                    cell.font = Font(color="0000FF", underline="single")
                except Exception:
                    # If hyperlink creation fails, just continue without it
                    pass

            if col_idx in currency_cols:
                cell.number_format = "#,##0.00"
            elif col_idx in integer_cols:
                cell.number_format = "0"
            elif col_idx in decimal_cols:
                cell.number_format = "0.0"

    # Totals row
    total_row = start_row + len(norm_items)
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=1).alignment = left
    for c in range(1, 21):
        ws.cell(row=total_row, column=c).border = thin_border

    ws.cell(row=total_row, column=9, value=int(totals["full_days"]))
    ws.cell(row=total_row, column=10, value=int(totals["half_days"]))
    ws.cell(row=total_row, column=12, value=totals["equivalent_days"]).number_format = (
        "0.0"
    )
    ws.cell(row=total_row, column=13, value=totals["ot_hours"]).number_format = "0.0"

    ws.cell(
        row=total_row, column=7, value=totals["base_daily_salary"]
    ).number_format = "#,##0.00"
    ws.cell(row=total_row, column=8, value=0).number_format = "#,##0.00"
    ws.cell(row=total_row, column=14, value=totals["daily_salary"]).number_format = (
        "#,##0.00"  # 日薪工資
    )
    ws.cell(
        row=total_row, column=15, value=totals["employee_mpf_amount"]
    ).number_format = "#,##0.00"
    ws.cell(
        row=total_row, column=16, value=totals["employer_mpf_amount"]
    ).number_format = "#,##0.00"

    ws.cell(row=total_row, column=17, value=totals["real_salary"]).number_format = (
        "#,##0.00"
    )
    ws.cell(row=total_row, column=17).fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    ws.cell(row=total_row, column=18, value=0).number_format = "#,##0.00"
    ws.cell(row=total_row, column=19, value=totals["total_payment"]).number_format = (
        "#,##0.00"
    )

    return ws


def generate_payslip_summary_xlsx(
    *,
    output_path: str,
    year: int,
    month: int,
    items: Sequence[ProjectSummaryItem] | Sequence[Dict[str, Any]],
) -> str:
    """Generate a monthly payslip summary Excel file (standalone)."""

    wb = Workbook()
    add_summary_sheet_to_workbook(
        wb,
        year=year,
        month=month,
        items=items,
        sheet_title="Summary",
        project_location="",
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


__all__ = [
    "ProjectSummaryItem",
    "generate_payslip_summary_xlsx",
    "add_summary_sheet_to_workbook",
]
